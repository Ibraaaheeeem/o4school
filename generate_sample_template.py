#!/usr/bin/env python3
"""
Generate sample Excel template for 4School bulk import
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# Create a new workbook
wb = openpyxl.Workbook()

# Remove default sheet
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# Define header style
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Example data style
example_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

# ============================================
# STUDENTS SHEET
# ============================================
students_sheet = wb.create_sheet("Students")

# Headers
students_headers = ["FirstName", "LastName", "MiddleName", "Gender", "DateOfBirth", "AdmissionNumber", "Address"]
students_sheet.append(students_headers)

# Style headers
for cell in students_sheet[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = border

# Add example rows
students_examples = [
    ["Ibrahim", "Musa", "Aliyu", "M", "15/05/2010", "ADM/2024/001", "123 Lagos Street, Kano"],
    ["Fatima", "Bello", "Amina", "F", "20/08/2011", "ADM/2024/002", "456 Abuja Road, Lagos"],
    ["Chioma", "Okonkwo", "", "F", "10/03/2012", "", "789 Port Harcourt Ave"]
]

for row_data in students_examples:
    students_sheet.append(row_data)
    for cell in students_sheet[students_sheet.max_row]:
        cell.fill = example_fill
        cell.border = border

# Set column widths
students_sheet.column_dimensions['A'].width = 15
students_sheet.column_dimensions['B'].width = 15
students_sheet.column_dimensions['C'].width = 15
students_sheet.column_dimensions['D'].width = 10
students_sheet.column_dimensions['E'].width = 15
students_sheet.column_dimensions['F'].width = 18
students_sheet.column_dimensions['G'].width = 30

# Add note
students_sheet.append([])
students_sheet.append(["NOTE: Gender must be 'M' or 'F'. Date format must be DD/MM/YYYY."])
note_cell = students_sheet[students_sheet.max_row][0]
note_cell.font = Font(italic=True, color="F59E0B", size=9)

# ============================================
# PARENTS SHEET
# ============================================
parents_sheet = wb.create_sheet("Parents")

# Headers
parents_headers = ["FirstName", "LastName", "Email", "PhoneNumber", "Address"]
parents_sheet.append(parents_headers)

# Style headers
for cell in parents_sheet[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = border

# Add example rows
parents_examples = [
    ["Aisha", "Musa", "aisha.musa@example.com", "08012345678", "123 Lagos Street, Kano"],
    ["Emeka", "Okonkwo", "emeka.okonkwo@example.com", "07098765432", "789 Port Harcourt Ave"],
    ["Zainab", "Bello", "zainab.bello@example.com", "08123456789", "456 Abuja Road, Lagos"]
]

for row_data in parents_examples:
    parents_sheet.append(row_data)
    for cell in parents_sheet[parents_sheet.max_row]:
        cell.fill = example_fill
        cell.border = border

# Set column widths
parents_sheet.column_dimensions['A'].width = 15
parents_sheet.column_dimensions['B'].width = 15
parents_sheet.column_dimensions['C'].width = 30
parents_sheet.column_dimensions['D'].width = 15
parents_sheet.column_dimensions['E'].width = 30

# Add note
parents_sheet.append([])
parents_sheet.append(["NOTE: Email and PhoneNumber must be unique. Email format will be validated."])
note_cell = parents_sheet[parents_sheet.max_row][0]
note_cell.font = Font(italic=True, color="F59E0B", size=9)

# ============================================
# STAFF SHEET
# ============================================
staff_sheet = wb.create_sheet("Staff")

# Headers
staff_headers = ["FirstName", "LastName", "Email", "PhoneNumber", "Designation", "DateOfHire"]
staff_sheet.append(staff_headers)

# Style headers
for cell in staff_sheet[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = border

# Add example rows
staff_examples = [
    ["Ngozi", "Adeyemi", "ngozi.adeyemi@school.com", "08034567890", "Teacher", "10/01/2022"],
    ["Yusuf", "Ibrahim", "yusuf.ibrahim@school.com", "07045678901", "Teacher", "15/03/2023"],
    ["Grace", "Eze", "grace.eze@school.com", "08156789012", "Librarian", ""]
]

for row_data in staff_examples:
    staff_sheet.append(row_data)
    for cell in staff_sheet[staff_sheet.max_row]:
        cell.fill = example_fill
        cell.border = border

# Set column widths
staff_sheet.column_dimensions['A'].width = 15
staff_sheet.column_dimensions['B'].width = 15
staff_sheet.column_dimensions['C'].width = 30
staff_sheet.column_dimensions['D'].width = 15
staff_sheet.column_dimensions['E'].width = 15
staff_sheet.column_dimensions['F'].width = 15

# Add note
staff_sheet.append([])
staff_sheet.append(["NOTE: Email and PhoneNumber must be unique. DateOfHire is optional (defaults to today)."])
note_cell = staff_sheet[staff_sheet.max_row][0]
note_cell.font = Font(italic=True, color="F59E0B", size=9)

# ============================================
# INSTRUCTIONS SHEET
# ============================================
instructions_sheet = wb.create_sheet("Instructions", 0)

# Title
instructions_sheet['A1'] = "4School Bulk Import Template"
instructions_sheet['A1'].font = Font(bold=True, size=16, color="667EEA")
instructions_sheet.merge_cells('A1:D1')

instructions_sheet['A3'] = "How to Use This Template:"
instructions_sheet['A3'].font = Font(bold=True, size=12)

instructions = [
    "",
    "1. This file contains three sheets: Students, Parents, and Staff",
    "2. Each sheet has example data in rows 2-4. You can delete these examples.",
    "3. Fill in your data starting from row 2 (row 1 contains headers)",
    "4. Do NOT change the header names or sheet names",
    "5. Required fields are marked in the format guide below",
    "",
    "IMPORTANT NOTES:",
    "• Date format must be DD/MM/YYYY (e.g., 15/05/2010)",
    "• Gender must be 'M' or 'F'",
    "• Email and PhoneNumber must be unique for Parents and Staff",
    "• AdmissionNumber must be unique for Students (if provided)",
    "• Maximum file size: 5MB",
    "",
    "STUDENTS - Required Fields:",
    "• FirstName, LastName, Gender, DateOfBirth",
    "",
    "PARENTS - Required Fields:",
    "• FirstName, LastName, Email, PhoneNumber",
    "",
    "STAFF - Required Fields:",
    "• FirstName, LastName, Email, PhoneNumber, Designation",
    "",
    "After filling in your data:",
    "1. Save this file",
    "2. Go to Community > Bulk Import in 4School",
    "3. Upload this file",
    "4. Review the validation results",
    "5. Confirm to import",
]

row = 4
for instruction in instructions:
    instructions_sheet[f'A{row}'] = instruction
    if instruction.startswith("•"):
        instructions_sheet[f'A{row}'].font = Font(size=10)
    elif instruction.endswith(":"):
        instructions_sheet[f'A{row}'].font = Font(bold=True, size=11)
    row += 1

instructions_sheet.column_dimensions['A'].width = 80

# Save the workbook
output_file = "4school_bulk_import_template.xlsx"
wb.save(output_file)

print(f"✅ Sample Excel template created: {output_file}")
print(f"📊 Sheets: {', '.join(wb.sheetnames)}")
print(f"📝 Students: {students_sheet.max_row - 1} example rows")
print(f"📝 Parents: {parents_sheet.max_row - 1} example rows")
print(f"📝 Staff: {staff_sheet.max_row - 1} example rows")
